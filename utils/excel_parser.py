"""Excel parsing utilities for loading evaluation criteria from Excel template."""
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Any
import openpyxl


def clean_text(value: Any) -> str:
    """Normalize text extracted from Excel cells.

    Args:
        value: Cell value from Excel (can be int, float, str, or None)

    Returns:
        Cleaned and normalized string
    """
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        if float(value).is_integer():
            value = int(value)
        return str(value)
    text = str(value)
    text = unicodedata.normalize("NFKC", text)

    # Replace special characters
    replacements = {
        "\u2013": "-",  # En dash
        "\u2014": "-",  # Em dash
        "\u2018": "'",  # Left single quote
        "\u2019": "'",  # Right single quote
        "\u201c": '"',  # Left double quote
        "\u201d": '"',  # Right double quote
        "\u2022": "•",  # Bullet
        "�": "'"        # Replacement character
    }
    for src, target in replacements.items():
        text = text.replace(src, target)

    # Normalize whitespace
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def slugify(*parts: str) -> str:
    """Generate a consistent key for rating items.

    Args:
        *parts: String parts to join and slugify

    Returns:
        URL-safe slug string
    """
    joined = "_".join(filter(None, parts))
    normalized = unicodedata.normalize("NFKD", joined)
    ascii_text = normalized.encode("ascii", "ignore").decode()
    ascii_text = ascii_text.lower()
    ascii_text = re.sub(r"[^a-z0-9]+", "_", ascii_text)
    ascii_text = re.sub(r"_+", "_", ascii_text).strip("_")
    return ascii_text or "criterion"


def merge_continued_sections(sections: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Combine sections that are continuations of the same group.

    For example, "PLATE WORK (Continued)" will be merged with "PLATE WORK".

    Args:
        sections: List of section dictionaries

    Returns:
        Merged list of sections
    """
    merged = {}
    order = []

    for section in sections:
        # Remove "(Continued)" or similar suffixes to get canonical name
        canonical = re.sub(r"\s*\(.*?\)\s*$", "", section["name"], flags=re.IGNORECASE).strip()

        if canonical not in merged:
            merged[canonical] = {
                "name": canonical,
                "raw_name": section["raw_name"],
                "subsections": []
            }
            order.append(canonical)
        else:
            # Update raw_name to include continuation info
            existing_raw = merged[canonical]["raw_name"]
            if section["raw_name"] not in existing_raw:
                merged[canonical]["raw_name"] = f"{existing_raw}, {section['raw_name']}"

        # Add subsections from this section
        merged[canonical]["subsections"].extend(section["subsections"])

    return [merged[name] for name in order]


def load_eval_criteria_from_excel(criteria_path: Path) -> List[Dict[str, Any]]:
    """Load evaluation criteria from the Excel template.

    Parses the "Eval. & Obser. Criteria" sheet to extract:
    - Section names (high-level categories like "PLATE WORK")
    - Subsections with max scores
    - Individual evaluation items with text and N/A option

    Args:
        criteria_path: Path to the Excel workbook

    Returns:
        List of section dictionaries with nested subsections and items
    """
    sections = []

    if not criteria_path.exists():
        return sections

    try:
        workbook = openpyxl.load_workbook(criteria_path, data_only=True)
    except Exception as exc:
        print(f"Failed to read evaluation template: {exc}")
        return sections

    sheet_name = "Eval. & Obser. Criteria"
    if sheet_name not in workbook.sheetnames:
        return sections

    sheet = workbook[sheet_name]
    current_section = None
    current_subsection = None
    last_item = None

    # Iterate through rows
    for row in sheet.iter_rows(min_row=1, max_col=11, values_only=True):
        col_a_raw = row[0]
        col_b_raw = row[1]
        col_c_raw = row[2]
        col_d_raw = row[3]
        col_e_raw = row[4]
        col_f_raw = row[5]
        others_raw = row[6:]

        # Clean all column values
        col_a = clean_text(col_a_raw)
        col_b = clean_text(col_b_raw)
        col_c = clean_text(col_c_raw)
        col_d = clean_text(col_d_raw)
        col_e = clean_text(col_e_raw)
        col_f = clean_text(col_f_raw)
        other_text = [clean_text(val) for val in others_raw]

        # Detect high-level section (e.g. PLATE WORK)
        # These typically have text in column E but not A, B, C
        if col_e and not col_a and not col_b and not col_c:
            label = col_e.strip()
            lowered = label.lower()
            # Skip if it's a scoring info line
            if not any(keyword in lowered for keyword in ("score", "out of", "pass or fail")):
                current_section = {
                    "name": label.title(),
                    "raw_name": label,
                    "subsections": []
                }
                sections.append(current_section)
                current_subsection = None
                last_item = None
            continue

        if not current_section:
            continue

        # Detect subsection row (contains Score or Pass/Fail info)
        info_line = " ".join(filter(None, [col_c, col_d, col_e, col_f] + other_text)).lower()
        if col_a and ("score" in info_line or "pass or fail" in info_line):
            # Extract max score from numeric cells
            numbers = []
            for value in [col_c_raw, col_d_raw, col_e_raw, col_f_raw] + list(others_raw):
                if isinstance(value, (int, float)):
                    if value:
                        numbers.append(int(value))
                elif isinstance(value, str):
                    match = re.search(r"(\d+)", value)
                    if match:
                        numbers.append(int(match.group(1)))

            max_score = max(numbers) if numbers else None
            current_subsection = {
                "name": col_a,
                "max_score": max_score,
                "items": []
            }
            current_section["subsections"].append(current_subsection)
            last_item = None
            continue

        if not current_subsection:
            continue

        # Detect evaluation item rows (text in column B)
        if col_b:
            text = col_b
            col_a_clean = col_a.strip()
            is_numeric_label = col_a_clean.replace(".", "", 1).isdigit()
            is_new_item = True

            # Check if this is a continuation of the previous item
            if last_item:
                if not col_a_clean:
                    is_new_item = False
                elif is_numeric_label:
                    first_char = text[:1]
                    if first_char and first_char.islower():
                        is_new_item = False
                    else:
                        # Check for connector words
                        lowered = text.lstrip().lower()
                        connector_prefixes = (
                            "and ", "or ", "nor ", "but ", "so ", "yet ",
                            "to ", "from ", "with ", "without ", "into ",
                            "onto ", "including ", "excluding "
                        )
                        if any(lowered.startswith(prefix) for prefix in connector_prefixes):
                            is_new_item = False
                elif not col_a_clean:
                    is_new_item = False
                else:
                    lowered = text.lstrip().lower()
                    if lowered and lowered[0].islower():
                        is_new_item = False

            if is_new_item:
                # Create new item
                key = slugify(
                    current_section["name"],
                    current_subsection["name"],
                    f"{len(current_subsection['items']) + 1:02d}"
                )
                allow_na = "n/a" in col_a.lower()
                item = {
                    "key": key,
                    "text": text,
                    "allow_na": allow_na
                }
                current_subsection["items"].append(item)
                last_item = item
            elif last_item:
                # Append to previous item
                joiner = "" if last_item["text"].endswith(("—", "-")) else " "
                last_item["text"] = f"{last_item['text']}{joiner}{text}".strip()
            continue

    return merge_continued_sections(sections)
